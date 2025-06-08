import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

# Directorios
crudos_dir = 'Crudos'
base_dir = 'Individuales'

# Verificación de carpeta base
if not os.path.isdir(base_dir):
    raise FileNotFoundError(f'La carpeta base "{base_dir}" no existe. Por favor, créala manualmente.')

# Obtener archivos crudos
archivos_crudos = [f for f in os.listdir(crudos_dir) if f.startswith("movimientos_") and f.endswith(".csv")]
meses_creados = []

# Procesar cada archivo mensual
for archivo in archivos_crudos:
    partes = archivo.replace('.csv', '').split('_')
    if len(partes) < 3:
        continue

    mes = partes[1]
    output_dir = os.path.join(base_dir, mes)

    if os.path.exists(output_dir):
        continue

    os.makedirs(output_dir)
    meses_creados.append(mes)

    input_path = os.path.join(crudos_dir, archivo)
    df = pd.read_csv(input_path)
    df["Fecha"] = pd.to_datetime(df["Fecha"], dayfirst=True, errors="coerce")  # Conversión segura de fecha

    nombres_unicos = df['Nombre'].unique()

    for nombre in nombres_unicos:
        df_persona = df[df['Nombre'] == nombre]
        nombre_archivo = f"{nombre.replace(' ', '_')}_{mes}.xlsx"
        ruta_archivo = os.path.join(output_dir, nombre_archivo)

        df_persona.to_excel(ruta_archivo, index=False)
        wb = load_workbook(ruta_archivo)
        ws = wb.active

        # Estilos de encabezado
        header_fill = PatternFill(start_color="177086", end_color="177086", fill_type="solid")
        header_border = Border(
            left=Side(style="thin", color="177086"),
            right=Side(style="thin", color="177086"),
            top=Side(style="thin", color="177086"),
            bottom=Side(style="thin", color="177086")
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            cell.border = header_border

        # Formato de columnas de pesos y dólares
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE

        for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
            for cell in row:
                cell.number_format = '"USD" #,##0.00'

        # Fórmulas de totales
        ultima_fila = ws.max_row + 1
        ws[f'E{ultima_fila}'] = f"=SUM(E2:E{ultima_fila - 1})"
        ws[f'F{ultima_fila}'] = f"=SUM(F2:F{ultima_fila - 1})"

        for col in ['E', 'F']:
            cell = ws[f'{col}{ultima_fila}']
            cell.font = Font(bold=True, color="444444")
            cell.fill = PatternFill(start_color="e3e8ea", end_color="e3e8ea", fill_type="solid")
            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(top=Side(style="medium", color="222222"))

        wb.save(ruta_archivo)

# Generar archivo combinado general
if meses_creados:
    print("Se generaron los siguientes meses:", ", ".join(meses_creados))

    salida_path = 'movimientos_totales_2025.csv'
    archivos = [f for f in os.listdir(crudos_dir) if f.startswith("movimientos_") and f.endswith(".csv")]

    dataframes = []
    for archivo in archivos:
        ruta = os.path.join(crudos_dir, archivo)
        df = pd.read_csv(ruta)
        df["Fecha"] = pd.to_datetime(df["Fecha"], dayfirst=True, errors="coerce")  # Asegurar formato uniforme
        dataframes.append(df)

    df_total = pd.concat(dataframes, ignore_index=True)
    df_total.to_csv(salida_path, index=False)

    print(f"Archivo combinado actualizado: {salida_path}")
else:
    print("Todos los meses ya estaban procesados. No se creó ningún archivo nuevo.")
